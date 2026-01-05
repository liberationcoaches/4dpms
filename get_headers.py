import openpyxl

wb = openpyxl.load_workbook('Execl/KRA Setting - 2025-26 (1).xlsx')

sheets_info = {
    'Organizational Dimension': 3,
    'Self Development': 3,
    'Developing Others': 3
}

for sheet_name, header_row in sheets_info.items():
    print(f"\n{'='*70}")
    print(f"Sheet: {sheet_name} (Row {header_row} headers)")
    print('='*70)
    ws = wb[sheet_name]
    
    row_data = [cell.value for cell in ws[header_row]]
    
    print("\nAll columns with headers:")
    headers = []
    for i, val in enumerate(row_data):
        if val and str(val).strip():
            col_letter = chr(65 + i) if i < 26 else f"{chr(65 + (i-26)//26)}{chr(65 + (i-26)%26)}"
            headers.append((col_letter, val))
            print(f"  Column {col_letter}: {val}")
    
    # Also check row 4 for sample data structure
    print("\nSample data row (row 4):")
    row4 = [cell.value for cell in ws[4]]
    for i, (col, header) in enumerate(headers):
        if i < len(row4) and row4[i] and str(row4[i]).strip() and not str(row4[i]).startswith('='):
            print(f"  {header}: {row4[i]}")

