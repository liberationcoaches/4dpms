import openpyxl
import json

wb = openpyxl.load_workbook('Execl/KRA Setting - 2025-26 (1).xlsx')
sheets = wb.sheetnames

print(f"Total sheets: {len(sheets)}\n")
print(f"Sheet names: {sheets}\n")

for sheet_name in sheets:
    print(f"{'='*60}")
    print(f"Sheet: {sheet_name}")
    print(f"{'='*60}")
    ws = wb[sheet_name]
    
    # Get headers from row 12
    row12 = [cell.value for cell in ws[12]]
    headers = []
    for i, val in enumerate(row12):
        if val:
            col_letter = chr(65+i) if i < 26 else f"A{chr(65+i-26)}"
            headers.append((col_letter, val))
    
    print(f"\nRow 12 Headers ({len(headers)} columns):")
    for col, header in headers:
        print(f"  {col}: {header}")
    
    # Check if structure is similar by looking at first data row
    print(f"\nChecking row 13-15 for data structure:")
    for row_num in range(13, 16):
        row_data = []
        for i, val in enumerate(row12[:20]):  # First 20 columns
            if val:
                cell_val = ws.cell(row=row_num, column=i+1).value
                if cell_val:
                    row_data.append((headers[i][1] if i < len(headers) else f"Col{i+1}", cell_val))
        if row_data:
            print(f"  Row {row_num}: {len(row_data)} fields")
            for header, val in row_data[:5]:  # First 5 fields
                print(f"    {header}: {val}")
    
    print("\n")

