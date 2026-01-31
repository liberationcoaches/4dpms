import openpyxl
import json

wb = openpyxl.load_workbook('Assets/Execl/Performance Apraisal Sheet - 2024-2025 -Align Sheet - Pune.xlsx')
sheet_name = 'Perfomance Appraisal Sheet'
ws = wb[sheet_name]

print(f"Analyzing sheet: {sheet_name}")
print(f"Max row: {ws.max_row}, Max column: {ws.max_column}\n")

# Analyze first 50 rows to understand structure
print("="*80)
print("STRUCTURE ANALYSIS")
print("="*80)

# Find header row (likely row with many non-empty cells)
header_row = None
for row_num in range(1, 20):
    non_empty = sum(1 for cell in ws[row_num] if cell.value)
    if non_empty > 5:
        header_row = row_num
        print(f"\nHeader row appears to be: Row {row_num} ({non_empty} non-empty cells)")
        break

if header_row:
    print(f"\nHeader values (first 30 columns):")
    headers = []
    for col_num in range(1, min(31, ws.max_column + 1)):
        cell = ws.cell(row=header_row, column=col_num)
        if cell.value:
            headers.append((col_num, cell.value))
            print(f"  Col {col_num}: {cell.value}")
    
    print(f"\nTotal header columns found: {len(headers)}")
    
    # Show sample data rows
    print(f"\nSample data rows (rows {header_row+1} to {header_row+5}):")
    for row_num in range(header_row + 1, min(header_row + 6, ws.max_row + 1)):
        print(f"\n  Row {row_num}:")
        for col_num, header_name in headers[:15]:  # First 15 columns
            cell = ws.cell(row=row_num, column=col_num)
            if cell.value:
                val = str(cell.value)[:50]  # First 50 chars
                print(f"    {header_name}: {val}")

# Check for merged cells
print(f"\n\nMerged cells:")
for merged in ws.merged_cells.ranges:
    print(f"  {merged}")

# Check for any patterns in column structure
print(f"\n\nColumn structure pattern:")
for col_num in range(1, min(31, ws.max_column + 1)):
    col_letter = openpyxl.utils.get_column_letter(col_num)
    non_empty = sum(1 for row in range(1, min(100, ws.max_row + 1)) if ws.cell(row=row, column=col_num).value)
    if non_empty > 0:
        print(f"  Column {col_letter} ({col_num}): {non_empty} non-empty cells in first 100 rows")
