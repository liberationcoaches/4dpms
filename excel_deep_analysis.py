import openpyxl
from openpyxl.utils import get_column_letter
import json

def analyze_excel_file(filepath, output_name):
    """Analyze a single Excel file and extract all relevant information"""
    print(f"\n{'='*100}")
    print(f"ANALYZING: {filepath}")
    print(f"{'='*100}")
    
    try:
        # Load workbook with data_only=False to get formulas
        wb_formulas = openpyxl.load_workbook(filepath, data_only=False)
        # Load workbook with data_only=True to get computed values
        wb_values = openpyxl.load_workbook(filepath, data_only=True)
        
        print(f"\nSheet names: {wb_formulas.sheetnames}")
        
        for sheet_name in wb_formulas.sheetnames:
            ws_f = wb_formulas[sheet_name]
            ws_v = wb_values[sheet_name]
            
            print(f"\n{'='*80}")
            print(f"SHEET: {sheet_name}")
            print(f"{'='*80}")
            print(f"Dimensions: {ws_f.max_row} rows x {ws_f.max_column} columns")
            
            # Find header rows (usually first few rows with text)
            print(f"\n--- HEADER STRUCTURE (First 15 rows) ---")
            for row_num in range(1, min(16, ws_f.max_row + 1)):
                row_data = []
                for col_num in range(1, min(30, ws_f.max_column + 1)):
                    cell_f = ws_f.cell(row=row_num, column=col_num)
                    cell_v = ws_v.cell(row=row_num, column=col_num)
                    if cell_f.value is not None:
                        col_letter = get_column_letter(col_num)
                        val = str(cell_f.value)[:50]
                        row_data.append(f"{col_letter}: {val}")
                if row_data:
                    print(f"  Row {row_num}: {', '.join(row_data[:8])}")
            
            # Find cells with formulas
            print(f"\n--- FORMULAS FOUND ---")
            formula_count = 0
            for row_num in range(1, ws_f.max_row + 1):
                for col_num in range(1, ws_f.max_column + 1):
                    cell = ws_f.cell(row=row_num, column=col_num)
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        col_letter = get_column_letter(col_num)
                        computed_value = ws_v.cell(row=row_num, column=col_num).value
                        print(f"  {col_letter}{row_num}: {cell.value}")
                        print(f"       Computed Value: {computed_value}")
                        formula_count += 1
                        if formula_count > 30:
                            print("  ... (truncated, more formulas exist)")
                            break
                if formula_count > 30:
                    break
            
            # Analyze data structure
            print(f"\n--- DATA SAMPLE (Rows 10-25) ---")
            for row_num in range(10, min(26, ws_v.max_row + 1)):
                row_data = []
                for col_num in range(1, min(20, ws_v.max_column + 1)):
                    cell = ws_v.cell(row=row_num, column=col_num)
                    if cell.value is not None:
                        col_letter = get_column_letter(col_num)
                        val = str(cell.value)[:40]
                        row_data.append(f"{col_letter}: {val}")
                if row_data:
                    print(f"  Row {row_num}: {', '.join(row_data[:6])}")
        
        wb_formulas.close()
        wb_values.close()
        
    except Exception as e:
        print(f"Error: {e}")

# Analyze all Excel files
files = [
    ('Assets/Execl/KRA Setting - 2025-26 (1).xlsx', 'KRA_Setting_Template'),
    ('Assets/Execl/Dispatch -  Chandrakant Karpe  24-25.xlsx', 'Dispatch_Chandrakant'),
    ('Assets/Execl/KRA Format_Sandipkumar Ashvinbhai Mevada.xlsx', 'KRA_Sandipkumar'),
    ('Assets/Execl/Performance Apraisal Sheet - 2024-2025 -Align Sheet - Pune.xlsx', 'Performance_Appraisal'),
]

for filepath, name in files:
    analyze_excel_file(filepath, name)
