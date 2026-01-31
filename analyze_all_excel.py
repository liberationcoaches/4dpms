import openpyxl
import json

# Excel files to analyze
excel_files = [
    'Assets/Execl/Dispatch -  Chandrakant Karpe  24-25.xlsx',
    'Assets/Execl/KRA Format_Sandipkumar Ashvinbhai Mevada.xlsx',
    'Assets/Execl/KRA Setting - 2025-26 (1).xlsx',
    'Assets/Execl/Performance Apraisal Sheet - 2024-2025 -Align Sheet - Pune.xlsx'
]

for excel_file in excel_files:
    print(f"\n{'='*80}")
    print(f"FILE: {excel_file}")
    print('='*80)
    
    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        
        print(f"\nSheet names: {list(wb.sheetnames)}")
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"\n--- Sheet: {sheet_name} ---")
            print(f"Dimensions: {ws.max_row} rows × {ws.max_column} columns")
            
            # Show first 10 rows
            print("\nFirst 10 rows:")
            for row_num in range(1, min(11, ws.max_row + 1)):
                row_data = []
                for col in range(1, min(21, ws.max_column + 1)):
                    cell = ws.cell(row=row_num, column=col)
                    if cell.value:
                        row_data.append((col, str(cell.value)[:40]))
                if row_data:
                    print(f"  Row {row_num}: {', '.join([f'C{c}: {v}' for c, v in row_data[:5]])}")
        
        wb.close()
    except Exception as e:
        print(f"Error reading file: {e}")
