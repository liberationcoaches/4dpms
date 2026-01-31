import openpyxl

wb = openpyxl.load_workbook('Assets/Execl/Performance Apraisal Sheet - 2024-2025 -Align Sheet - Pune.xlsx')
ws = wb['Perfomance Appraisal Sheet']

print("="*80)
print("DETAILED COLUMN STRUCTURE")
print("="*80)

# Check rows 7-9 to understand the header structure
for row_num in range(7, 10):
    print(f"\nRow {row_num}:")
    for col_num in range(1, 26):
        cell = ws.cell(row=row_num, column=col_num)
        if cell.value:
            col_letter = openpyxl.utils.get_column_letter(col_num)
            print(f"  {col_letter}{row_num}: {cell.value}")

# Check a sample data row to see what values are in each column
print("\n" + "="*80)
print("SAMPLE DATA ROW (Row 10 - Mr. Sagar Gavhane)")
print("="*80)
for col_num in range(1, 26):
    cell = ws.cell(row=10, column=col_num)
    if cell.value:
        col_letter = openpyxl.utils.get_column_letter(col_num)
        print(f"  {col_letter}10: {cell.value}")

# Check what sub-columns exist for R1, R2, R3, R4
print("\n" + "="*80)
print("SUB-COLUMNS FOR EACH REVIEW PERIOD")
print("="*80)

# R1 is in columns D-G (4-7)
print("\nR1 Score columns (D-G):")
for col_num in range(4, 8):
    cell = ws.cell(row=8, column=col_num)
    col_letter = openpyxl.utils.get_column_letter(col_num)
    print(f"  {col_letter}8: {cell.value if cell.value else '(empty)'}")

# R2 is in columns H-K (8-11)
print("\nR2 Score columns (H-K):")
for col_num in range(8, 12):
    cell = ws.cell(row=8, column=col_num)
    col_letter = openpyxl.utils.get_column_letter(col_num)
    print(f"  {col_letter}8: {cell.value if cell.value else '(empty)'}")

# R3 is in columns L-O (12-15)
print("\nR3 Score columns (L-O):")
for col_num in range(12, 16):
    cell = ws.cell(row=8, column=col_num)
    col_letter = openpyxl.utils.get_column_letter(col_num)
    print(f"  {col_letter}8: {cell.value if cell.value else '(empty)'}")

# R4 is in columns P-S (16-19)
print("\nR4 Score columns (P-S):")
for col_num in range(16, 20):
    cell = ws.cell(row=8, column=col_num)
    col_letter = openpyxl.utils.get_column_letter(col_num)
    print(f"  {col_letter}8: {cell.value if cell.value else '(empty)'}")

# Check row 10 values in detail
print("\n" + "="*80)
print("ROW 10 DETAILED VALUES")
print("="*80)
for col_num in range(1, 26):
    cell = ws.cell(row=10, column=col_num)
    if cell.value is not None:
        col_letter = openpyxl.utils.get_column_letter(col_num)
        print(f"  {col_letter}10: {cell.value} (type: {type(cell.value).__name__})")
